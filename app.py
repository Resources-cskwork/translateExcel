from flask import Flask, request, send_file, render_template_string, jsonify
import io
import openpyxl
from openpyxl.cell.cell import MergedCell
import logging
import requests
import json
import copy
import random
import concurrent.futures
from functools import lru_cache
from collections import defaultdict
import threading
import queue
import time
from datetime import datetime
from werkzeug.utils import secure_filename

# Constants and Configuration
MAX_FILE_SIZE = 50 * 1024 * 1024  # 50MB limit
ALLOWED_EXTENSIONS = {'xlsx', 'xls'}
MAX_RETRIES = 3
OLLAMA_API_URL = 'http://localhost:11434/api/generate'
OLLAMA_MODEL = 'gemma2:latest'

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    handlers=[
        logging.StreamHandler(),
        logging.FileHandler('translation_app.log')
    ]
)
logger = logging.getLogger(__name__)

def allowed_file(filename):
    """Check if file extension is allowed."""
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

def check_ollama_health():
    """Check if Ollama service is running and model is available."""
    try:
        response = requests.post(
            OLLAMA_API_URL,
            json={'model': OLLAMA_MODEL, 'prompt': 'test', 'stream': False},
            timeout=1000
        )
        return response.status_code == 200
    except Exception as e:
        logger.error(f"Ollama health check failed: {str(e)}")
        return False

app = Flask(__name__)

@app.before_request
def before_request():
    """Middleware to check service health before each request."""
    if not check_ollama_health():
        return jsonify({'error': 'Translation service is currently unavailable'}), 503

# Thread-safe translation cache
translation_cache = {}
cache_lock = threading.Lock()

# Translation queue for batch processing
translation_queue = queue.Queue()
BATCH_SIZE = 10

@lru_cache(maxsize=1000)
def cached_translate(text, context_hash=None):
    """Cache translation results for repeated text."""
    return translate_with_context(text, context_hash)

def batch_translate(texts, context):
    """Translate multiple texts in a single API call."""
    if not texts:
        return []
    
    try:
        # First attempt: batch translation
        combined_text = "\n---SPLIT---\n".join(texts)
        combined_translation = translate_with_context(combined_text, context)
        
        if not combined_translation:
            logger.warning("Batch translation returned empty result, trying individual translations")
            return retry_failed_translations(texts, None, context)
        
        translations = combined_translation.split("\n---SPLIT---\n")
        
        # If we got fewer translations than input texts, retry the missing ones
        if len(translations) != len(texts):
            logger.warning(f"Translation count mismatch (got {len(translations)}, expected {len(texts)}), retrying failed translations")
            return retry_failed_translations(texts, translations, context)
            
        return translations
    except Exception as e:
        logger.error(f"Batch translation failed: {str(e)}, trying individual translations")
        return retry_failed_translations(texts, None, context)

def retry_failed_translations(original_texts, partial_translations=None, context=None):
    """Retry failed translations individually."""
    final_translations = []
    
    # If we have partial translations, use them where available
    if partial_translations:
        valid_translations = partial_translations[:len(original_texts)]
    else:
        valid_translations = [None] * len(original_texts)
    
    for idx, (orig_text, trans) in enumerate(zip(original_texts, valid_translations)):
        if trans:  # Use existing translation if available
            final_translations.append(trans)
            continue
            
        # Try individual translation up to 3 times
        success = False
        for attempt in range(3):
            try:
                individual_translation = translate_with_context(orig_text, context)
                if individual_translation and individual_translation.strip():
                    final_translations.append(individual_translation)
                    success = True
                    break
            except Exception as e:
                logger.error(f"Individual translation attempt {attempt + 1} failed for text: {orig_text[:100]}... Error: {str(e)}")
                time.sleep(1)  # Wait before retry
        
        if not success:
            logger.warning(f"All individual translation attempts failed for text: {orig_text[:100]}...")
            final_translations.append(orig_text)  # Use original if all retries fail
    
    return final_translations

def process_cell_batch(batch, context):
    """Process a batch of cells in parallel."""
    unique_texts = {}
    for cell, text in batch:
        if text:
            unique_texts[text] = unique_texts.get(text, []) + [cell]
    
    # Translate unique texts
    translations = batch_translate(list(unique_texts.keys()), context)
    
    # Apply translations back to cells
    results = []
    for text, translation in zip(unique_texts.keys(), translations):
        for cell in unique_texts[text]:
            results.append((cell, translation))
    
    return results

def analyze_excel_structure(workbook, max_items=50):
    """Analyze Excel file structure and content to build context."""
    context = {
        'headers': set(),         # Column headers
        'row_headers': set(),     # Row headers/labels
        'dates': set(),          # Date formats found
        'times': set(),          # Time formats found
        'formulas': set(),       # Formula patterns
        'currencies': set(),      # Currency formats
        'bullet_types': set(),    # Types of bullets used
        'repeated_terms': {},     # Frequently occurring terms
        'cell_formats': set()     # Special cell formats
    }
    
    try:
        # Count total cells to process
        total_cells = sum(sheet.max_row * sheet.max_column for sheet in workbook)
        if total_cells > 10000:  # Limit for very large files
            logger.warning(f"Large file detected ({total_cells} cells). Using sampling for analysis.")
            sampling_rate = max(0.1, 10000 / total_cells)  # Sample at most 10000 cells
        else:
            sampling_rate = 1.0

        for sheet in workbook:
            # Analyze headers (first row only)
            for cell in list(sheet[1])[:max_items]:
                if cell.value:
                    context['headers'].add(str(cell.value))
            
            # Analyze first column (limited number of row headers)
            for row in list(sheet.iter_rows())[:max_items]:
                if row[0].value:
                    context['row_headers'].add(str(row[0].value))
            
            # Sample cells for pattern analysis
            rows = list(sheet.iter_rows())
            if sampling_rate < 1.0:
                rows = random.sample(rows, int(len(rows) * sampling_rate))
            
            for row in rows:
                for cell in row:
                    if cell.value and random.random() < sampling_rate:
                        try:
                            value = str(cell.value)
                            
                            # Limit collection sizes
                            if len(context['formulas']) < max_items and isinstance(value, str) and value.startswith('='):
                                context['formulas'].add(value[:100])  # Limit formula length
                            
                            if len(context['bullet_types']) < 10 and isinstance(value, str) and value.strip().startswith(('•', '-', '□', '○', '◆')):
                                context['bullet_types'].add(value[0])
                            
                            if cell.number_format:
                                if len(context['dates']) < 10 and ('y' in cell.number_format.lower() or 'm' in cell.number_format.lower() or 'd' in cell.number_format.lower()):
                                    context['dates'].add(cell.number_format)
                                
                                if len(context['times']) < 10 and ('h' in cell.number_format.lower() or ':' in cell.number_format):
                                    context['times'].add(cell.number_format)
                                
                                if len(context['currencies']) < 10 and any(curr in cell.number_format for curr in ['$', '￦', '¥', '€', 'RM']):
                                    context['currencies'].add(cell.number_format)
                                
                                if len(context['cell_formats']) < max_items and cell.number_format != 'General':
                                    context['cell_formats'].add(cell.number_format)
                            
                            # Count repeated terms (limit length and quantity)
                            if isinstance(value, str) and len(value) > 1 and len(value) <= 100:
                                if len(context['repeated_terms']) < max_items:
                                    context['repeated_terms'][value] = context['repeated_terms'].get(value, 0) + 1

                        except Exception as e:
                            logger.warning(f"Error processing cell: {str(e)}")
                            continue

        # Filter and limit repeated terms
        context['repeated_terms'] = dict(sorted(
            [(k, v) for k, v in context['repeated_terms'].items() if v > 1],
            key=lambda x: x[1],
            reverse=True
        )[:max_items])

        # Convert sets to lists and limit sizes
        for key in context:
            if isinstance(context[key], set):
                context[key] = list(context[key])[:max_items]

        return context

    except Exception as e:
        logger.error(f"Error in analyze_excel_structure: {str(e)}")
        return {key: [] if isinstance(value, set) else {} if isinstance(value, dict) else value 
                for key, value in context.items()}

def translate_text(text, context=None):
    if not text or not isinstance(text, str):
        return text

    try:
        logger.info(f"Translating text: {text}")
        
        # Skip translation for formulas, dates, and pure numbers
        if text.startswith('='):
            return text
        if text.replace('.', '').replace('/', '').replace('-', '').isdigit():
            return text
            
        # Check if the text matches any special patterns from context
        if context:
            # Preserve headers if they're being reused
            if text in context['headers']:
                logger.info(f"Found header pattern: {text}")
            
            # Preserve row labels if they're being reused
            if text in context['row_headers']:
                logger.info(f"Found row header pattern: {text}")
            
            # Preserve bullet points
            if any(text.startswith(bullet) for bullet in context['bullet_types']):
                # Only translate the text after the bullet
                bullet = text[0]
                remaining_text = text[1:].strip()
                translated_remaining = translate_with_context(remaining_text, context)
                return f"{bullet} {translated_remaining}"
        
        return translate_with_context(text, context)
            
    except Exception as e:
        logger.error(f"Translation error: {str(e)}")
        return text

def translate_with_context(text, context=None):
    """Translate text with context awareness and retries."""
    if not text:
        return text
        
    original_text = text
    for attempt in range(MAX_RETRIES):
        try:
            context_str = ""
            if context:
                # Limit context size
                headers = ', '.join(context['headers'][:10])
                common_terms = ', '.join(k for k, v in list(context['repeated_terms'].items())[:10])
                date_formats = ', '.join(context['dates'][:5])
                time_formats = ', '.join(context['times'][:5])
                currency_formats = ', '.join(context['currencies'][:5])
                bullet_types = ', '.join(context['bullet_types'][:5])

                context_str = f"""Document Context:
- Common Headers: {headers[:200]}
- Frequent Terms: {common_terms[:200]}
- Date Formats: {date_formats[:100]}
- Time Formats: {time_formats[:100]}
- Currency Formats: {currency_formats[:100]}
- Bullet Types: {bullet_types[:50]}
"""

            # Limit prompt size
            prompt = f"""You are a professional Korean to English translator specializing in business documents and Excel spreadsheets.
Please translate the following Korean text to English:

Text to translate: {text[:600]}

{context_str[:1000]}

Translation requirements:
1. Maintain all numbers, special characters, and formatting exactly as in the original
2. Keep all file paths, URLs, and email addresses in their original form
3. Preserve all technical terms, company names, and proper nouns
4. For currency amounts, keep the original numbers and currency symbols
5. If there are abbreviations or technical terms you're uncertain about, keep them in Korean
6. Translate consistently with other similar terms in the document

Important: Provide ONLY the English translation without any explanations or notes.
"""

            # Make the API call to Ollama with timeout
            response = requests.post(
                OLLAMA_API_URL,
                json={
                    'model': OLLAMA_MODEL,
                    'prompt': prompt,
                    'stream': False,
                    'temperature': 0.3,
                    'top_p': 0.9
                },
                timeout=30
            )
            
            if response.status_code != 200:
                logger.error(f"Translation API error: {response.status_code}")
                continue
                
            result = response.json()
            if not result.get('response'):
                logger.error("Empty translation response")
                continue
                
            return result['response'].strip()
            
        except Exception as e:
            logger.error(f"Translation attempt {attempt + 1} failed: {str(e)}")
            if attempt == MAX_RETRIES - 1:
                logger.warning(f"All translation attempts failed, returning original text: {original_text[:100]}...")
                return original_text
            continue
        
        time.sleep(1)  # Wait before retry
    
    return original_text  # Return original if all retries failed

def is_numeric_string(text):
    """Check if a string contains only numbers, decimals, and basic number formatting."""
    if not isinstance(text, str):
        return False
    # Remove common number formatting characters
    cleaned = text.replace(',', '').replace('$', '').replace('%', '').strip()
    try:
        float(cleaned)
        return True
    except ValueError:
        return False

def process_excel(file):
    logger.info("Starting Excel processing")
    try:
        workbook = openpyxl.load_workbook(file)
        context = analyze_excel_structure(workbook)
        
        # Create output workbook
        new_wb = openpyxl.Workbook()
        new_wb.remove(new_wb.active)  # Remove default sheet
        
        # Process sheets in parallel
        with concurrent.futures.ThreadPoolExecutor(max_workers=min(4, len(workbook.sheetnames))) as executor:
            futures = []
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                new_sheet = new_wb.create_sheet(title=sheet_name)
                
                # Copy sheet properties
                if sheet.sheet_format:
                    new_sheet.sheet_format = copy.copy(sheet.sheet_format)
                if sheet.sheet_properties:
                    new_sheet.sheet_properties = copy.copy(sheet.sheet_properties)
                
                # Submit sheet processing task
                futures.append(executor.submit(process_sheet, sheet, new_sheet, context))
            
            # Wait for all sheets to complete
            for future in concurrent.futures.as_completed(futures):
                try:
                    future.result()
                except Exception as e:
                    logger.error(f"Error processing sheet: {str(e)}")
        
        # Save to buffer
        buffer = io.BytesIO()
        new_wb.save(buffer)
        buffer.seek(0)
        return buffer
        
    except Exception as e:
        logger.error(f"Error in process_excel: {str(e)}")
        raise

def process_sheet(sheet, new_sheet, context):
    """Process a single sheet with batched cell processing."""
    try:
        # Copy sheet dimensions
        for key, value in sheet.column_dimensions.items():
            new_sheet.column_dimensions[key] = copy.copy(value)
        for key, value in sheet.row_dimensions.items():
            new_sheet.row_dimensions[key] = copy.copy(value)
        
        # Copy merged cells
        if sheet.merged_cells:
            new_sheet.merged_cells = copy.copy(sheet.merged_cells)
        
        # Collect cells for batch processing
        cells_to_process = []
        for row in sheet.iter_rows():
            for cell in row:
                if isinstance(cell, MergedCell):
                    continue
                
                new_cell = new_sheet.cell(row=cell.row, column=cell.column)
                
                # Copy cell formatting
                if cell.has_style:
                    new_cell.font = copy.copy(cell.font)
                    new_cell.border = copy.copy(cell.border)
                    new_cell.fill = copy.copy(cell.fill)
                    new_cell.number_format = copy.copy(cell.number_format)
                    new_cell.protection = copy.copy(cell.protection)
                    new_cell.alignment = copy.copy(cell.alignment)
                
                # Add to translation batch if needed
                if cell.value and isinstance(cell.value, str) and not is_numeric_string(cell.value):
                    cells_to_process.append((new_cell, cell.value))
                else:
                    new_cell.value = cell.value
        
        # Process cells in batches
        for i in range(0, len(cells_to_process), BATCH_SIZE):
            batch = cells_to_process[i:i + BATCH_SIZE]
            results = process_cell_batch(batch, context)
            
            # Apply translations
            for cell, translation in results:
                cell.value = translation
        
        logger.info(f"Completed processing sheet: {sheet.title}")
        
    except Exception as e:
        logger.error(f"Error processing sheet {sheet.title}: {str(e)}")
        raise

# HTML template for the upload page
HTML_TEMPLATE = """
<!DOCTYPE html>
<html>
<head>
    <title>Excel Korean to English Translator</title>
    <style>
        body {
            font-family: Arial, sans-serif;
            max-width: 800px;
            margin: 0 auto;
            padding: 20px;
            text-align: center;
        }
        .container {
            background-color: #f9f9f9;
            padding: 20px;
            border-radius: 8px;
            box-shadow: 0 2px 4px rgba(0,0,0,0.1);
        }
        h1 {
            color: #333;
        }
        .upload-form {
            margin: 20px 0;
        }
        .message {
            margin: 10px 0;
            padding: 10px;
            border-radius: 4px;
        }
        .success {
            background-color: #d4edda;
            color: #155724;
        }
        .error {
            background-color: #f8d7da;
            color: #721c24;
        }
    </style>
</head>
<body>
    <div class="container">
        <h1>Excel Korean to English Translator</h1>
        <p>Upload an Excel file to translate its content from Korean to English</p>
        
        <form class="upload-form" action="/" method="post" enctype="multipart/form-data">
            <input type="file" name="file" accept=".xlsx" required>
            <br><br>
            <input type="submit" value="Translate">
        </form>
        
        {% if error %}
        <div class="message error">
            {{ error }}
        </div>
        {% endif %}
    </div>
</body>
</html>
"""

@app.route('/', methods=['GET', 'POST'])
def upload_file():
    if request.method == 'POST':
        try:
            # Check if file was uploaded
            if 'file' not in request.files:
                return jsonify({'error': 'No file uploaded'}), 400
            
            file = request.files['file']
            if file.filename == '':
                return jsonify({'error': 'No file selected'}), 400
            
            # Validate file type
            if not allowed_file(file.filename):
                return jsonify({'error': 'Invalid file type. Only Excel files (.xlsx, .xls) are allowed'}), 400
            
            # Check file size
            file_content = file.read()
            if len(file_content) > MAX_FILE_SIZE:
                return jsonify({'error': 'File size exceeds 50MB limit'}), 400
            
            # Process file with retries
            for attempt in range(MAX_RETRIES):
                try:
                    buffer = process_excel(io.BytesIO(file_content))
                    
                    # Create response
                    output = io.BytesIO(buffer.getvalue())
                    output.seek(0)
                    
                    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
                    filename = f"translated_{timestamp}_{secure_filename(file.filename)}"
                    
                    return send_file(
                        output,
                        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                        as_attachment=True,
                        download_name=filename
                    )
                
                except Exception as e:
                    if attempt == MAX_RETRIES - 1:
                        logger.error(f"Failed to process file after {MAX_RETRIES} attempts: {str(e)}")
                        return jsonify({'error': 'Failed to process file. Please try again later.'}), 500
                    logger.warning(f"Attempt {attempt + 1} failed, retrying...")
                    time.sleep(1)  # Wait before retry
            
        except Exception as e:
            logger.error(f"Upload error: {str(e)}")
            return jsonify({'error': 'An unexpected error occurred'}), 500
    
    return render_template_string(HTML_TEMPLATE)

if __name__ == '__main__':
    logger.info("Starting Flask application...")
    app.run(debug=True, host='0.0.0.0', port=5001)
